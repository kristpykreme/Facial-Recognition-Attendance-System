''''
Based on original code by Anirban Kar: https://github.com/thecodacus/Face-Recognition    

Developed by Marcelo Rovai - MJRoBot.org @ 21Feb18  

'''
### -------- import for CAM -------- ###
import cv2
import numpy as np
import os

### -------- import for EXCEl -------- ###
from openpyxl import Workbook
from openpyxl import load_workbook
import os.path


### -------- EXCEL SETUP -------- ###

def Taking_Attendance(date):
# check for existing attendence list, create one if does not exist
    if os.path.isfile('Attendance.xlsx'):
        book = load_workbook('Attendance.xlsx')
        print("\nLoading existing attendance sheet...")
    else:
        book = Workbook()
        print("\nCreating new attendance sheet...")

    sheet = book.create_sheet()
    sheet.title = '{}'.format(date)

    # format namelist 

    name_list = (
        ['Index', 'Name', 'Attendance'],
        [1, 'Brendon'],
        [2, 'Raymond'],
        [3, 'Tzi Seong'],
        )

    for rows in name_list: 

        sheet.append(rows)

    ### -------- CAM SETUP -------- ###
        
    recognizer = cv2.face.LBPHFaceRecognizer_create()
    recognizer.read('trainer/trainer.yml')
    cascadePath = "haarcascade_frontalface_default.xml"
    faceCascade = cv2.CascadeClassifier(cascadePath);

    font = cv2.FONT_HERSHEY_SIMPLEX

    id = 0

    # namelist for cam
    names = ['None','Brendon (1)','Raymond (2)','Tzi Seong (3)']

    # Initialize and start realtime video capture
    cam = cv2.VideoCapture(0)
    cam.set(3, 640) # set video widht
    cam.set(4, 480) # set video height

    # Define min window size to be recognized as a face
    minW = 0.1*cam.get(3)
    minH = 0.1*cam.get(4)

    while True:

        ret, img =cam.read()
        img = cv2.flip(img, 1)

        gray = cv2.cvtColor(img,cv2.COLOR_BGR2GRAY)

        faces = faceCascade.detectMultiScale( 
            gray,
            scaleFactor = 1.2,
            minNeighbors = 5,
            minSize = (int(minW), int(minH)),
           )

        for(x,y,w,h) in faces:

            cv2.rectangle(img, (x,y), (x+w,y+h), (0,255,0), 2)

            id, confidence = recognizer.predict(gray[y:y+h,x:x+w])

            # Check if confidence is less then 100 ==> "0" is perfect match 
            if (confidence < 100):
                sheet.cell(row = id + 1, column = 3).value = 1
                id = names[id]
                confidence = "  {0}%".format(round(100 - confidence))
            else:
                id = "unknown"
                confidence = "  {0}%".format(round(100 - confidence))
            
            cv2.putText(img, str(id), (x+5,y-5), font, 1, (255,255,255), 2)
            cv2.putText(img, str(confidence), (x+5,y+h-5), font, 1, (255,255,0), 1)  
        
        cv2.imshow('camera',img)

        k = cv2.waitKey(10) & 0xff # Press 'ESC' for exiting video
        if k == 27:
            break

    ### -------- CHECK WHO IS ABSENT -------- ###
        
    absent_list = []

    for x in range(2, len(name_list)+1):
        present = sheet['C{}'.format(x)]
        if present.value != 1:
            name = sheet['B{}'.format(x)]
            absent_list.append(name.value)
            
    ### -------- CLEANUP CAM -------- ###
    print("\n---------------------------------")
    print("\n---------Exiting Program---------")
    cam.release()
    cv2.destroyAllWindows()

    ### -------- CLEANUP EXCEl -------- ###
    book.save('Attendance.xlsx')
    book.close()

    print("\nAbsentee(s): {} ".format(len(absent_list)))
    for y in absent_list:
        print(y)
    print("\nAttendance for {} saved.".format(date))
